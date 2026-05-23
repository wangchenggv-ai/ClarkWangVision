# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import math

OUTPUT = r"C:\Users\wangc\Downloads\ClarkWangVision\04-供应链\feishu-setup\inventory-system\heat-map-tool\暑期备货模型.xlsx"

# ── 原始热力图数据 ──────────────────────────────────────────
HEATMAP = {
    0.00:  [152, 9, 36, 24, 22, 3, 2, 1, 4],
   -0.25:  [43, 10, 20, 14, 13, 3, 3, 3, 4],
   -0.50:  [70, 6, 46, 21, 11, 3, 2, 4, 3],
   -0.75:  [92, 6, 55, 28, 15, 3, 2, 4, 4],
   -1.00:  [110, 13, 74, 44, 18, 1, 1, 3, 4],
   -1.25:  [116, 15, 78, 22, 19, 2, 2, 2, 4],
   -1.50:  [91, 6, 82, 38, 25, 0, 1, 4, 3],
   -1.75:  [85, 5, 55, 39, 26, 3, 3, 3, 2],
   -2.00:  [65, 9, 60, 40, 26, 2, 3, 4, 3],
   -2.25:  [19, 2, 16, 4, 3, 3, 4, 3, 2],
   -2.50:  [9, 2, 11, 11, 4, 4, 4, 4, 4],
   -2.75:  [17, 3, 17, 7, 5, 3, 3, 4, 2],
   -3.00:  [14, 4, 7, 3, 4, 1, 3, 2, 2],
   -3.25:  [5, 1, 9, 3, 6, 2, 2, 2, 4],
   -3.50:  [5, 2, 7, 1, 5, 3, 2, 3, 4],
   -3.75:  [4, 1, 3, 2, 6, 2, 4, 4, 2],
   -4.00:  [1, 2, 2, 5, 1, 4, 1, 3, 1],
   -4.25:  [2, 1, 3, 1, 4, 2, 3, 4, 4],
   -4.50:  [3, 2, 2, 4, 3, 3, 1, 4, 2],
   -4.75:  [4, 2, 3, 1, 4, 1, 2, 4, 3],
   -5.00:  [2, 2, 3, 2, 3, 4, 2, 4, 4],
   -5.25:  [3, 0, 0, 4, 2, 3, 4, 4, 3],
   -5.50:  [2, 2, 2, 3, 4, 3, 4, 5, 4],
   -5.75:  [3, 2, 2, 3, 2, 2, 4, 1, 3],
   -6.00:  [4, 1, 1, 4, 3, 3, 3, 2, 4],
}
CYL_VALS = [0.00, -0.25, -0.50, -0.75, -1.00, -1.25, -1.50, -1.75, -2.00]
SPH_VALS = sorted(HEATMAP.keys(), reverse=True)  # 0 → -6
HIST_TOTAL = sum(v for row in HEATMAP.values() for v in row)  # 2473

# ── 辅助颜色 / 样式 ────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def bold(size=10, color="000000"):
    return Font(bold=True, size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

YELLOW_FILL  = fill("FFF2CC")  # 参数输入格
BLUE_FILL    = fill("1F4E79")   # 表头深蓝
LBLUE_FILL   = fill("BDD7EE")   # 表头浅蓝
GREEN_FILL   = fill("E2EFDA")   # 小标题
GRAY_FILL    = fill("F2F2F2")
RED_FILL     = fill("FF4C4C")
ORANGE_FILL  = fill("FFA040")
LYELLOW_FILL = fill("FFFF99")
WHITE_FILL   = fill("FFFFFF")

# ── 计算核心数据 ───────────────────────────────────────────
def has_xianpian(sph, cyl):
    return sph >= -4.0 and cyl >= -1.0

# 每个 SKU 的基础预测（不含安全库存，参数格控制）
def base_forecast(sph, cyl):
    idx = CYL_VALS.index(cyl)
    return HEATMAP[sph][idx] * 12000 / HIST_TOTAL

# ── 构建 SKU 列表 ──────────────────────────────────────────
skus = []
for sph in SPH_VALS:
    for cyl in CYL_VALS:
        idx = CYL_VALS.index(cyl)
        hist = HEATMAP[sph][idx]
        fc = base_forecast(sph, cyl)
        xp = has_xianpian(sph, cyl)
        skus.append(dict(sph=sph, cyl=cyl, hist=hist, fc=fc, xp=xp))

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════
# TAB 1: 参数设置
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "参数设置"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 28
ws1.column_dimensions["D"].width = 16

def param_row(ws, row, label, val, unit="", note=""):
    ws.cell(row, 1, label).alignment = right()
    ws.cell(row, 1).font = Font(size=10)
    c = ws.cell(row, 2, val)
    c.fill = YELLOW_FILL
    c.alignment = center()
    c.font = Font(bold=True, size=11, color="1F4E79")
    c.border = thin_border()
    ws.cell(row, 3, unit).font = Font(size=9, color="666666")
    ws.cell(row, 3).alignment = Alignment(vertical="center")
    if note:
        ws.cell(row, 4, note).font = Font(size=9, color="888888", italic=True)

# 标题
ws1.merge_cells("A1:D1")
c = ws1["A1"]
c.value = "暑期备货模型 — 参数设置"
c.fill = fill("1F4E79")
c.font = Font(bold=True, size=14, color="FFFFFF")
c.alignment = center()
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:D2")
ws1["A2"].value = "黄色格可直接修改，其余各Tab自动更新"
ws1["A2"].font = Font(size=9, color="888888", italic=True)
ws1["A2"].alignment = center()

# 分组标题
def section(ws, row, text):
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1, text)
    c.fill = GREEN_FILL
    c.font = Font(bold=True, size=10, color="375623")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20

section(ws1, 3, "▌ 需求参数")
param_row(ws1, 4,  "暑假预计总销量",      12000, "片（=6000副×2）",  "可调整")
param_row(ws1, 5,  "历史热力图总量",       2473,  "片（自动，勿改）", "用于计算放大系数")
param_row(ws1, 6,  "放大系数",             "=B4/B5", "×",             "自动计算")

section(ws1, 7, "▌ 现片参数")
param_row(ws1, 8,  "现片 MOQ（最低起订量）", 100,  "片/SKU",          "每个度数最少订100片")
param_row(ws1, 9,  "现片下单门槛",            50,  "片（预测需求）",  "低于此值改车房片")
param_row(ws1, 10, "现片交货周期",             8,  "周",              "约2个月，需提前下单")
param_row(ws1, 11, "现片价格系数",             1,  "（相对）",        "车房片=3，用于成本估算")

section(ws1, 12, "▌ 车房片参数")
param_row(ws1, 13, "车房片交货周期",           1,  "周",              "按订单生产")
param_row(ws1, 14, "车房片价格系数",            3,  "（相对）",        "是现片的3倍")

section(ws1, 15, "▌ 库存风险参数")
param_row(ws1, 16, "现片残值率",            0.5,   "（淡季消化50%）", "超备现片可低价处理")
param_row(ws1, 17, "安全库存系数（A类）",   0.20,  "预测量×(1+系数)", "预测≥200片的SKU")
param_row(ws1, 18, "安全库存系数（B类）",   0.10,  "预测量×(1+系数)", "预测100–199片的SKU")
param_row(ws1, 19, "安全库存系数（C类）",   0.00,  "无额外安全库存",  "预测50–99片的SKU，卡MOQ")

section(ws1, 20, "▌ 排期参数")
param_row(ws1, 21, "暑假开始日期",   "2026-07-01", "",               "决定现片截止下单日")
param_row(ws1, 22, "现片截止下单日", "=TEXT(DATE(YEAR(B21),MONTH(B21),DAY(B21))-B10*7,\"YYYY-MM-DD\")",
          "", "暑假开始 - 现片交期")

ws1.row_dimensions[22].height = 18

# 备注框
ws1.merge_cells("A24:D28")
c = ws1["A24"]
c.value = ("【决策规则说明】\n"
           "① 预测需求 ≥ 200片（A类）：现片，含20%安全库存\n"
           "② 预测需求 100–199片（B类）：现片，含10%安全库存\n"
           "③ 预测需求 50–99片（C类）：现片，MOQ=100，不加安全库存\n"
           "④ 预测需求 < 50片，或超出现片度数范围：全部车房片")
c.font = Font(size=9, color="444444")
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
c.fill = fill("EBF3FB")
c.border = thin_border()

# ════════════════════════════════════════════════════════════
# TAB 2: 需求预测矩阵
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("需求预测矩阵")
ws2.sheet_view.showGridLines = False

# 标题
ws2.merge_cells("A1:K1")
c = ws2["A1"]
c.value = "暑假需求预测热力图（片）"
c.fill = fill("1F4E79")
c.font = Font(bold=True, size=13, color="FFFFFF")
c.alignment = center()
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:K2")
ws2["A2"].value = "= 历史销量 × (总目标 / 历史总量)，含安全库存系数；颜色越深需求越高"
ws2["A2"].font = Font(size=9, color="666666", italic=True)
ws2["A2"].alignment = center()

# CYL 表头
ws2.cell(4, 1, "SPH \\ CYL").fill = BLUE_FILL
ws2.cell(4, 1).font = Font(bold=True, size=10, color="FFFFFF")
ws2.cell(4, 1).alignment = center()
ws2.column_dimensions["A"].width = 9

for ci, cyl in enumerate(CYL_VALS):
    col = ci + 2
    c = ws2.cell(4, col, f"{cyl:.2f}")
    c.fill = BLUE_FILL
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.alignment = center()
    ws2.column_dimensions[get_column_letter(col)].width = 8

# 总计列
ws2.cell(4, 11, "行合计").fill = BLUE_FILL
ws2.cell(4, 11).font = Font(bold=True, size=10, color="FFFFFF")
ws2.cell(4, 11).alignment = center()
ws2.column_dimensions["K"].width = 9

ws2.row_dimensions[4].height = 22

# 图例行
ws2.merge_cells("A3:K3")
ws2["A3"].value = "■ ≥400  ■ 200–399  ■ 100–199  ■ 50–99  □ <50（车房片）"
ws2["A3"].font = Font(size=9, color="444444")
ws2["A3"].alignment = center()

# 颜色方案
def fc_fill(val):
    if val >= 400:   return fill("C00000")   # 深红
    elif val >= 200: return fill("FF6B6B")   # 红
    elif val >= 100: return fill("FFB347")   # 橙
    elif val >= 50:  return fill("FFD966")   # 黄
    else:            return fill("F2F2F2")   # 灰

# 安全库存系数
def safety(fc):
    if fc >= 200: return 0.20
    elif fc >= 100: return 0.10
    else: return 0.00

# 数据行
for ri, sph in enumerate(SPH_VALS):
    row = ri + 5
    ws2.row_dimensions[row].height = 18
    # SPH 标签
    c = ws2.cell(row, 1, f"{sph:.2f}")
    c.fill = LBLUE_FILL
    c.font = Font(bold=True, size=10, color="1F4E79")
    c.alignment = center()
    row_sum = 0
    for ci, cyl in enumerate(CYL_VALS):
        col = ci + 2
        hist = HEATMAP[sph][ci]
        base_fc = hist * 12000 / HIST_TOTAL
        sc = safety(base_fc)
        adj_fc = base_fc * (1 + sc)
        val = round(adj_fc)
        row_sum += val
        c = ws2.cell(row, col, val if hist > 0 else "")
        c.fill = fc_fill(val) if hist > 0 else fill("FAFAFA")
        c.font = Font(size=9, bold=(val >= 200),
                      color="FFFFFF" if val >= 400 else "000000")
        c.alignment = center()
        c.border = thin_border()
    # 行合计
    c = ws2.cell(row, 11, row_sum)
    c.fill = LBLUE_FILL
    c.font = Font(bold=True, size=9, color="1F4E79")
    c.alignment = center()

# 列合计行
total_row = len(SPH_VALS) + 5
ws2.cell(total_row, 1, "列合计").fill = LBLUE_FILL
ws2.cell(total_row, 1).font = Font(bold=True, size=9, color="1F4E79")
ws2.cell(total_row, 1).alignment = center()
grand = 0
for ci, cyl in enumerate(CYL_VALS):
    col = ci + 2
    col_sum = sum(
        round(HEATMAP[sph][ci] * 12000 / HIST_TOTAL * (1 + safety(HEATMAP[sph][ci] * 12000 / HIST_TOTAL)))
        for sph in SPH_VALS
    )
    grand += col_sum
    c = ws2.cell(total_row, col, col_sum)
    c.fill = LBLUE_FILL
    c.font = Font(bold=True, size=9, color="1F4E79")
    c.alignment = center()
ws2.cell(total_row, 11, grand).fill = fill("1F4E79")
ws2.cell(total_row, 11).font = Font(bold=True, size=10, color="FFFFFF")
ws2.cell(total_row, 11).alignment = center()

# ════════════════════════════════════════════════════════════
# TAB 3: SKU 决策表
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("SKU决策表")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A5"

ws3.merge_cells("A1:K1")
c = ws3["A1"]
c.value = "SKU 备货决策表（225个SKU全览）"
c.fill = fill("1F4E79")
c.font = Font(bold=True, size=13, color="FFFFFF")
c.alignment = center()
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A2:K2")
ws3["A2"].value = "绿色行=现片下单；灰色行=车房片按需；★=超备超过50片，需关注"
ws3["A2"].font = Font(size=9, color="444444", italic=True)
ws3["A2"].alignment = center()

headers = ["SPH", "CYL", "历史\n销量", "基础\n预测", "安全\n系数", "含安全\n预测",
           "有现片?", "决策", "现片\n订量", "超备\n数量", "备注"]
col_widths = [8, 8, 8, 9, 8, 9, 8, 10, 9, 9, 18]

ws3.merge_cells("A3:K3")
ws3["A3"].value = "— 按预测需求降序排列 —"
ws3["A3"].font = Font(size=9, color="888888", italic=True)
ws3["A3"].alignment = center()

for ci, (h, w) in enumerate(zip(headers, col_widths)):
    col = ci + 1
    c = ws3.cell(4, col, h)
    c.fill = BLUE_FILL
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.alignment = center()
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.row_dimensions[4].height = 30

# 构建决策行，按预测降序
THRESHOLD = 50
rows_data = []
for s in skus:
    base_fc = s['fc']
    sc = safety(base_fc)
    adj_fc = base_fc * (1 + sc)
    use_xp = s['xp'] and adj_fc >= THRESHOLD
    if use_xp:
        order_qty = math.ceil(adj_fc / 100) * 100
        order_qty = max(order_qty, 100)
        overage = order_qty - adj_fc
    else:
        order_qty = 0
        overage = 0
    rows_data.append(dict(
        sph=s['sph'], cyl=s['cyl'], hist=s['hist'],
        base_fc=round(base_fc, 1), sc=sc, adj_fc=round(adj_fc, 1),
        xp=s['xp'], use_xp=use_xp,
        order_qty=order_qty, overage=round(overage, 0),
    ))

rows_data.sort(key=lambda x: -x['adj_fc'])

for ri, d in enumerate(rows_data):
    row = ri + 5
    ws3.row_dimensions[row].height = 16
    row_fill = fill("E8F5E9") if d['use_xp'] else fill("F8F8F8")

    def wcell(col, val, fmt=None, bold_=False, color_="000000", align_=None):
        c = ws3.cell(row, col, val)
        c.fill = row_fill
        c.font = Font(size=9, bold=bold_, color=color_)
        c.border = thin_border()
        c.alignment = align_ or center()
        if fmt:
            c.number_format = fmt
        return c

    wcell(1, d['sph'], "0.00", bold_=True, color_="1F4E79")
    wcell(2, d['cyl'], "0.00", bold_=True, color_="1F4E79")
    wcell(3, d['hist'])
    wcell(4, d['base_fc'], "0.0")
    wcell(5, d['sc'], "0%")
    wcell(6, d['adj_fc'], "0.0")
    wcell(7, "Y" if d['xp'] else "N",
          bold_=d['xp'], color_="006400" if d['xp'] else "999999")
    decision_text = "现片" if d['use_xp'] else ("车房片" if d['hist'] > 0 else "无需备")
    dcell = ws3.cell(row, 8, decision_text)
    dcell.fill = fill("C8E6C9") if d['use_xp'] else (fill("E3F2FD") if d['hist'] > 0 else fill("F5F5F5"))
    dcell.font = Font(size=9, bold=d['use_xp'],
                      color="1B5E20" if d['use_xp'] else ("1565C0" if d['hist'] > 0 else "AAAAAA"))
    dcell.border = thin_border()
    dcell.alignment = center()

    if d['use_xp']:
        wcell(9, d['order_qty'], bold_=True, color_="1B5E20")
    else:
        c = ws3.cell(row, 9, "—")
        c.fill = row_fill; c.border = thin_border(); c.alignment = center()
        c.font = Font(size=9, color="BBBBBB")

    if d['use_xp']:
        ov = d['overage']
        flag = "★ 超备多" if ov > 50 else ""
        wcell(10, round(ov), color_="C62828" if ov > 50 else "444444")
        wcell(11, flag, color_="C62828", bold_=bool(flag))
    else:
        for col in [10, 11]:
            c = ws3.cell(row, col, "—")
            c.fill = row_fill; c.border = thin_border(); c.alignment = center()
            c.font = Font(size=9, color="BBBBBB")

# 合计行
total_row3 = len(rows_data) + 5
ws3.cell(total_row3, 1, "合计").fill = fill("1F4E79")
ws3.cell(total_row3, 1).font = Font(bold=True, size=10, color="FFFFFF")
ws3.cell(total_row3, 1).alignment = center()
ws3.cell(total_row3, 9, sum(d['order_qty'] for d in rows_data)).fill = fill("1F4E79")
ws3.cell(total_row3, 9).font = Font(bold=True, size=10, color="FFFFFF")
ws3.cell(total_row3, 9).alignment = center()
ws3.cell(total_row3, 10, round(sum(d['overage'] for d in rows_data if d['use_xp']))).fill = fill("1F4E79")
ws3.cell(total_row3, 10).font = Font(bold=True, size=10, color="FFFFFF")
ws3.cell(total_row3, 10).alignment = center()

# ════════════════════════════════════════════════════════════
# TAB 4: 订货清单 & 成本分析
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("订货清单")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "A5"

ws4.merge_cells("A1:H1")
c = ws4["A1"]
c.value = "现片订货清单 & 成本对比"
c.fill = fill("1F4E79")
c.font = Font(bold=True, size=13, color="FFFFFF")
c.alignment = center()
ws4.row_dimensions[1].height = 30

# 成本汇总框
xp_skus = [d for d in rows_data if d['use_xp']]
cf_skus  = [d for d in rows_data if not d['use_xp'] and d['hist'] > 0]
total_xp_order = sum(d['order_qty'] for d in xp_skus)
total_xp_fc    = sum(d['adj_fc'] for d in xp_skus)
total_cf_fc    = sum(d['adj_fc'] for d in cf_skus)
xp_overage     = total_xp_order - total_xp_fc

# P = 车房片单价参考值(相对), 现片=1/3P
# 现片方案成本 = order_qty×(P/3) - overage×残值×(P/3)
# 全车房片成本 = adj_fc×P

P = 3  # 相对单位，车房片=3

def cost_xianpian(d, residual=0.5):
    net = d['order_qty'] * 1 - (d['order_qty'] - d['adj_fc']) * 1 * residual
    return net

def cost_chefang(d):
    return d['adj_fc'] * P

total_cost_hybrid  = sum(cost_xianpian(d) for d in xp_skus) + sum(cost_chefang(d) for d in cf_skus)
total_cost_all_cf  = sum(cost_chefang(d) for d in (xp_skus + cf_skus))
saving = total_cost_all_cf - total_cost_hybrid
saving_pct = saving / total_cost_all_cf * 100

summary_data = [
    ("现片 SKU 数量",         f"{len(xp_skus)} 个",       ""),
    ("现片合计订量",           f"{total_xp_order:,} 片",   ""),
    ("现片预计销出",           f"{round(total_xp_fc):,} 片",""),
    ("现片超备量",             f"{round(xp_overage):,} 片","淡季消化50%"),
    ("车房片 SKU 数量",        f"{len(cf_skus)} 个",       ""),
    ("车房片预计需求",          f"{round(total_cf_fc):,} 片","按订单生产"),
    ("——","——","——"),
    ("相对成本：混合方案",      f"{total_cost_hybrid:,.0f} P",""),
    ("相对成本：全车房片方案",  f"{total_cost_all_cf:,.0f} P",""),
    ("混合方案节省",           f"{saving:,.0f} P  ({saving_pct:.1f}%)", "★"),
]

ws4.merge_cells("A2:D2")
ws4["A2"].value = "▌ 方案汇总"
ws4["A2"].fill = GREEN_FILL
ws4["A2"].font = Font(bold=True, size=10, color="375623")
ws4["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

for ri, (label, val, note) in enumerate(summary_data):
    row = ri + 3
    if label == "——":
        ws4.merge_cells(f"A{row}:D{row}")
        ws4.cell(row, 1).fill = fill("EEEEEE")
        continue
    lc = ws4.cell(row, 1, label)
    lc.fill = GRAY_FILL; lc.font = Font(size=9); lc.border = thin_border()
    lc.alignment = Alignment(horizontal="right", vertical="center")
    vc = ws4.cell(row, 2, val)
    vc.fill = YELLOW_FILL if "节省" in label else WHITE_FILL
    vc.font = Font(size=10, bold=("节省" in label or "订量" in label), color="C00000" if "节省" in label else "000000")
    vc.border = thin_border()
    vc.alignment = center()
    ws4.cell(row, 3, note).font = Font(size=9, color="888888", italic=True)

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 16

# 现片订单明细
order_header_row = len(summary_data) + 4
ws4.merge_cells(f"A{order_header_row}:H{order_header_row}")
ws4.cell(order_header_row, 1, "▌ 现片订货明细（按订量降序）").fill = GREEN_FILL
ws4.cell(order_header_row, 1).font = Font(bold=True, size=10, color="375623")
ws4.cell(order_header_row, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)

oh = order_header_row + 1
order_cols = ["序号", "SPH", "CYL", "含安全库存\n预测(片)", "现片\n订量", "超备量", "相对成本\n（现片方案）", "相对成本\n（全车房）"]
order_widths = [6, 8, 8, 14, 10, 10, 16, 14]
for ci, (h, w) in enumerate(zip(order_cols, order_widths)):
    c = ws4.cell(oh, ci+1, h)
    c.fill = BLUE_FILL; c.font = Font(bold=True, size=9, color="FFFFFF")
    c.alignment = center(); c.border = thin_border()
    ws4.column_dimensions[get_column_letter(ci+1)].width = w
ws4.row_dimensions[oh].height = 28

xp_sorted = sorted(xp_skus, key=lambda x: -x['order_qty'])
for ri, d in enumerate(xp_sorted):
    row = oh + 1 + ri
    ws4.row_dimensions[row].height = 16
    vals = [ri+1, f"{d['sph']:.2f}", f"{d['cyl']:.2f}",
            round(d['adj_fc'], 0), d['order_qty'],
            round(d['overage']),
            round(cost_xianpian(d), 1), round(cost_chefang(d), 1)]
    for ci, v in enumerate(vals):
        c = ws4.cell(row, ci+1, v)
        c.fill = fill("F9FBF9") if ri % 2 == 0 else WHITE_FILL
        c.font = Font(size=9, bold=(ci in [4]))
        c.border = thin_border()
        c.alignment = center()

# 合计
total_row4 = oh + 1 + len(xp_sorted)
ws4.cell(total_row4, 1, "合计").fill = fill("1F4E79"); ws4.cell(total_row4, 1).font = Font(bold=True, size=10, color="FFFFFF"); ws4.cell(total_row4, 1).alignment = center()
ws4.cell(total_row4, 5, total_xp_order).fill = fill("1F4E79"); ws4.cell(total_row4, 5).font = Font(bold=True, size=10, color="FFFFFF"); ws4.cell(total_row4, 5).alignment = center()
ws4.cell(total_row4, 7, round(sum(cost_xianpian(d) for d in xp_sorted), 1)).fill = fill("1F4E79"); ws4.cell(total_row4, 7).font = Font(bold=True, size=10, color="FFFFFF"); ws4.cell(total_row4, 7).alignment = center()
ws4.cell(total_row4, 8, round(sum(cost_chefang(d) for d in xp_sorted), 1)).fill = fill("1F4E79"); ws4.cell(total_row4, 8).font = Font(bold=True, size=10, color="FFFFFF"); ws4.cell(total_row4, 8).alignment = center()

# ════════════════════════════════════════════════════════════
# TAB 5: 排期
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("备货排期")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 16
ws5.column_dimensions["B"].width = 20
ws5.column_dimensions["C"].width = 42
ws5.column_dimensions["D"].width = 24

ws5.merge_cells("A1:D1")
c = ws5["A1"]
c.value = "暑期备货排期（倒排）"
c.fill = fill("1F4E79")
c.font = Font(bold=True, size=13, color="FFFFFF")
c.alignment = center()
ws5.row_dimensions[1].height = 30

ws5.merge_cells("A2:D2")
ws5["A2"].value = "关键节点：现片截止下单日 = 暑假开始 - 8周；车房片滚动按需"
ws5["A2"].font = Font(size=9, color="666666", italic=True)
ws5["A2"].alignment = center()

timeline = [
    ("今天",         "2026-05-19", "确认现片SKU清单 + 与工厂对接产能", "★ 紧急：工厂暑假排产已开始"),
    ("T-10周",       "2026-05-26", "现片订单正式提交工厂，锁定产线", "距截止1周，务必完成"),
    ("T-8周",        "2026-06-09", "现片下单截止（最迟）",           "⚠ 超过此日期无法暑假前到货"),
    ("T-6周",        "2026-06-23", "跟催工厂生产进度，确认交期",     ""),
    ("T-3周",        "2026-07-14", "现片到货验收入库",              "距暑假高峰3周"),
    ("T-2周",        "2026-07-21", "车房片第一批预下单（热门度数）", "覆盖头2周订单"),
    ("暑假高峰",      "2026-07-01\n~08-31", "车房片滚动：周二下单→次周到货", "每周滚动补货"),
    ("暑假结束",      "2026-09-01", "盘点现片剩余库存",             "超备片进入淡季消化"),
    ("淡季",         "2026-09 起", "现片剩余按低价走货，避免积压",  ""),
]

hdr_cols = ["时间节点", "日期", "行动事项", "备注"]
for ci, h in enumerate(hdr_cols):
    c = ws5.cell(3, ci+1, h)
    c.fill = BLUE_FILL; c.font = Font(bold=True, size=10, color="FFFFFF")
    c.alignment = center(); c.border = thin_border()
ws5.row_dimensions[3].height = 22

urgent_rows = {4, 5, 6}  # 今天、T-10周、T-8周
for ri, (node, date, action, note) in enumerate(timeline):
    row = ri + 4
    ws5.row_dimensions[row].height = 22
    is_urgent = row in urgent_rows
    row_fill = fill("FFF2CC") if is_urgent else (fill("F9F9F9") if ri % 2 == 0 else WHITE_FILL)
    for ci, val in enumerate([node, date, action, note]):
        c = ws5.cell(row, ci+1, val)
        c.fill = row_fill
        c.font = Font(size=9, bold=is_urgent, color="C00000" if is_urgent else "333333")
        c.border = thin_border()
        c.alignment = Alignment(horizontal="left" if ci >= 2 else "center",
                                vertical="center", wrap_text=True)

# 现片SKU清单附在排期后
ws5.merge_cells(f"A{len(timeline)+6}:D{len(timeline)+6}")
ws5.cell(len(timeline)+6, 1, "-> 现片订单明细请见[订货清单]Tab").fill = fill("EBF3FB")
ws5.cell(len(timeline)+6, 1).font = Font(size=10, bold=True, color="1565C0")
ws5.cell(len(timeline)+6, 1).alignment = center()

# ── 保存 ──────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Done: {OUTPUT}")
print(f"现片SKU: {len(xp_skus)}个，订量: {total_xp_order}片")
print(f"车房片SKU: {len(cf_skus)}个，预测: {round(total_cf_fc)}片")
print(f"节省: {saving:.0f}P ({saving_pct:.1f}%)")
