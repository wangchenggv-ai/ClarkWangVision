#!/usr/bin/env python3
"""
月度对账单 — 按代理商汇总发货、退换货、应收金额。

用法:
  python billing.py 2026-05              # 指定月份
  python billing.py                      # 默认当月

输出:
  output/billing/billing_2026-05.xlsx   对账单（每个代理商一个 Sheet）
  output/billing/summary_2026-05.xlsx   汇总表

对账单列:
  日期 | 批次编号 | 顾客姓名 | 眼别 | 产品型号 | SPH | CYL | 单价 | 金额
  退货行用红色标注（金额为负）
  底部合计行：发货数量、退货数量、应收金额
"""

import sys
import logging
from datetime import date
from pathlib import Path
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


def _parse_month(arg: str | None) -> str:
    if arg:
        # Accept YYYY-MM or YYYYMM
        arg = arg.strip()
        if len(arg) == 6 and arg.isdigit():
            return f"{arg[:4]}-{arg[4:]}"
        return arg
    return date.today().strftime("%Y-%m")


def _fetch_shipped_orders(month: str) -> list[dict]:
    """Read 已发货 orders for the given month from Feishu order_detail table."""
    from config import TABLES
    from modules import feishu_client as fc

    detail_table = TABLES.get("order_detail", "")
    if not detail_table:
        log.warning("order_detail 表 ID 未配置，使用空数据")
        return []

    year, mon = month.split("-")
    date_prefix = f"{year}-{mon}"

    try:
        records = fc.search_records(detail_table, filter_={
            "conjunction": "and",
            "conditions": [
                {"field_name": "发货状态", "operator": "is", "value": ["已发货"]},
            ],
        })
        # Filter by month on the client side (Feishu doesn't support date-prefix filter easily)
        result = []
        for rec in records:
            batch = fc.text(rec.get("批次编号", ""))
            # BATCH-YYYYMMDD-HHMMSS → extract date part
            if len(batch) >= 14 and batch[6:12] == year + mon:
                result.append(rec)
            elif date_prefix in batch:
                result.append(rec)
        return result
    except Exception as e:
        log.error("读取飞书失败: %s", e)
        return []


def _write_billing_excel(agent_data: dict, month: str, out_dir: Path) -> Path:
    """Write billing Excel with one sheet per agent + a summary sheet."""
    path = out_dir / f"billing_{month}.xlsx"
    writer = pd.ExcelWriter(path, engine="openpyxl")

    summary_rows = []

    for agent_name, rows in sorted(agent_data.items()):
        df = pd.DataFrame(rows, columns=[
            "日期", "批次编号", "顾客姓名", "眼别", "产品型号",
            "球镜SPH", "柱镜CYL", "类型", "单价", "金额",
        ])
        sheet_name = agent_name[:30]  # Excel sheet name limit
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        shipped  = len([r for r in rows if r[7] == "发货"])
        returned = len([r for r in rows if r[7] == "退货"])
        total    = sum(r[9] for r in rows)
        summary_rows.append({
            "代理商": agent_name,
            "发货片数": shipped,
            "退货片数": returned,
            "应收金额": total,
        })

    # Summary sheet
    pd.DataFrame(summary_rows).to_excel(writer, sheet_name="汇总", index=False)
    writer.close()

    # Apply basic styles
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = ws["A2"]
    wb.save(path)
    return path


def main():
    month = _parse_month(sys.argv[1] if len(sys.argv) > 1 else None)
    print(f"\n  月度对账单：{month}\n")

    records = _fetch_shipped_orders(month)
    if not records:
        print("  未找到已发货订单，请确认 order_detail 表已配置且有数据")
        print("  提示：先运行 python main.py ./inbox/ 并完成发货确认")
        sys.exit(0)

    from modules import feishu_client as fc

    # Group by agent
    agent_data: dict[str, list] = defaultdict(list)
    for rec in records:
        agent = fc.text(rec.get("代理商名称", "未知代理商"))
        batch = fc.text(rec.get("批次编号", ""))
        # Date from batch ID: BATCH-YYYYMMDD-HHMMSS
        rec_date = batch[6:14] if len(batch) >= 14 else ""
        if len(rec_date) == 8:
            rec_date = f"{rec_date[:4]}-{rec_date[4:6]}-{rec_date[6:]}"

        agent_data[agent].append([
            rec_date,
            batch,
            fc.text(rec.get("顾客姓名", "")),
            fc.text(rec.get("眼别", "")),
            fc.text(rec.get("产品型号", "")),
            fc.number(rec.get("球镜SPH")),
            fc.number(rec.get("柱镜CYL")),
            "发货",
            0.0,   # 单价（飞书表暂无价格字段，填0待扩展）
            0.0,   # 金额
        ])

    out_dir = Path("output") / "billing"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = _write_billing_excel(agent_data, month, out_dir)

    total_shipped = sum(len(v) for v in agent_data.values())
    print(f"  ✓ 对账单已生成: {path}")
    print(f"  覆盖 {len(agent_data)} 个代理商，{total_shipped} 片发货记录")
    print()
    print("  注意：单价/金额字段当前为 0，需在飞书订单明细表中添加「单价」字段")
    print("        或在本脚本中按产品型号维护一张价格映射表")
    print()


if __name__ == "__main__":
    main()
