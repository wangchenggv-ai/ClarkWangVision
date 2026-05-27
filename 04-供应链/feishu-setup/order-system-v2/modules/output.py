"""
Generate output Excel files.

labels.xlsx  — 配货单，按货位排序，仓库拣货用
factory.xlsx — 排产单，按 SKU+SPH+CYL 聚合，工厂排产用
errors.xlsx  — 异常记录（可选，仅当有异常时生成）
"""

import logging
from pathlib import Path
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# Column widths (approximate character width)
_LABEL_WIDTHS = {
    "序列号": 8, "货位": 12, "顾客姓名": 12, "眼别": 6,
    "产品型号": 14, "球镜SPH": 8, "柱镜CYL": 8, "轴位AXIS": 8,
    "镜片码": 20, "验真网址": 40, "代理商": 14, "终端门店": 14,
    "联系人": 10, "联系电话": 14, "收货地址": 30, "备注": 20,
}
_FACTORY_WIDTHS = {
    "序号": 6, "产品型号": 14, "球镜SPH": 8, "柱镜CYL": 8,
    "轴位AXIS": 8, "数量": 6, "序列号": 8, "关联代理商": 30, "备注": 20,
}


def _apply_styles(ws, header_row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for cell in ws[header_row]:
        cell.fill   = header_fill
        cell.font   = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border    = border
            cell.alignment = center

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _set_col_widths(ws, widths: dict) -> None:
    for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        header = col[0].value or ""
        w = widths.get(header, 14)
        ws.column_dimensions[get_column_letter(i)].width = w


def write_labels(label_list: list[dict], out_dir: Path) -> Path:
    """Generate 配货单 sorted by bin_location → serial_no → customer."""
    rows = sorted(
        label_list,
        key=lambda r: (
            r.get("bin_location") or "ZZZ",
            r.get("serial_no") or "",
            r.get("customer_name") or "",
            r.get("eye") or "",
        ),
    )

    data = []
    for r in rows:
        data.append({
            "序列号":    r.get("serial_no", ""),
            "货位":      r.get("bin_location", ""),
            "顾客姓名":  r.get("customer_name", ""),
            "眼别":      r.get("eye", ""),
            "产品型号":  r.get("product_sku", ""),
            "球镜SPH":   r.get("sph", ""),
            "柱镜CYL":   r.get("cyl", ""),
            "轴位AXIS":  r.get("axis", "") or "",
            "镜片码":    r.get("lens_code", ""),
            "验真网址":  r.get("verify_url", ""),
            "代理商":    r.get("agent_name", ""),
            "终端门店":  r.get("store_name", ""),
            "联系人":    r.get("contact", ""),
            "联系电话":  r.get("phone", ""),
            "收货地址":  r.get("address", ""),
            "备注":      r.get("note", ""),
        })

    path = out_dir / "labels.xlsx"
    df = pd.DataFrame(data)
    df.to_excel(path, index=False, sheet_name="配货单")
    wb = load_workbook(path)
    ws = wb.active
    _apply_styles(ws)
    _set_col_widths(ws, _LABEL_WIDTHS)
    wb.save(path)
    log.info("labels.xlsx: %d 行 → %s", len(data), path)
    return path


def write_factory(factory_list: list[dict], out_dir: Path) -> Path | None:
    """Generate 排产单 aggregated by SKU+SPH+CYL."""
    if not factory_list:
        log.info("排产单: 0 行，跳过")
        return None

    # Aggregate by (product_sku, sph, cyl, axis)
    from collections import defaultdict
    agg: dict[tuple, dict] = defaultdict(lambda: {
        "qty": 0, "agents": set(), "serial_no": "", "note": ""
    })
    for r in factory_list:
        key = (
            r.get("product_sku", ""),
            r.get("sph"),
            r.get("cyl"),
            r.get("axis"),
        )
        agg[key]["qty"] += 1
        agent = r.get("agent_name") or r.get("agent_code", "")
        if agent:
            agg[key]["agents"].add(agent)
        if not agg[key]["serial_no"]:
            agg[key]["serial_no"] = r.get("serial_no", "")
        if r.get("note"):
            agg[key]["note"] = r["note"]

    # Sort: product_sku → sph → cyl
    sorted_keys = sorted(
        agg.keys(),
        key=lambda k: (k[0], k[1] if k[1] is not None else 0, k[2] if k[2] is not None else 0),
    )

    data = []
    for i, key in enumerate(sorted_keys, start=1):
        sku, sph, cyl, axis = key
        entry = agg[key]
        data.append({
            "序号":     i,
            "产品型号": sku,
            "球镜SPH":  sph,
            "柱镜CYL":  cyl,
            "轴位AXIS": axis or "",
            "数量":     entry["qty"],
            "序列号":   entry["serial_no"],
            "关联代理商": "、".join(sorted(entry["agents"])),
            "备注":     entry["note"],
        })

    path = out_dir / "factory.xlsx"
    df = pd.DataFrame(data)
    df.to_excel(path, index=False, sheet_name="排产单")
    wb = load_workbook(path)
    ws = wb.active
    _apply_styles(ws)
    _set_col_widths(ws, _FACTORY_WIDTHS)
    wb.save(path)
    log.info("factory.xlsx: %d 行 → %s", len(data), path)
    return path


def write_errors(errors: list[str], match_errors: list[dict], out_dir: Path) -> Path | None:
    """Write parse and match errors to errors.xlsx if any exist."""
    rows = []
    for e in errors:
        rows.append({"来源": "解析", "文件/行": "", "问题": e})
    for r in match_errors:
        for err in r.get("_match_errors", []):
            rows.append({
                "来源":   "匹配",
                "文件/行": f"{r.get('_file', '')} 行{r.get('_row', '')}",
                "问题":   err,
            })
    if not rows:
        return None
    path = out_dir / "errors.xlsx"
    pd.DataFrame(rows).to_excel(path, index=False)
    log.warning("errors.xlsx: %d 条问题 → %s", len(rows), path)
    return path


def prepare_output_dir() -> Path:
    today = date.today().isoformat()
    out = Path("output") / today
    out.mkdir(parents=True, exist_ok=True)
    return out
