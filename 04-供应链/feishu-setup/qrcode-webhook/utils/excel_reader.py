import os
import sys
import uuid
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from config import Config

# Maps internal Python keys → Excel column header strings (Chinese)
COLUMN_MAP = {
    "order_id":        "订单号",
    "customer_name":   "客户姓名",
    "left_sph":        "左眼球镜",
    "left_cyl":        "左眼柱镜",
    "right_sph":       "右眼球镜",
    "right_cyl":       "右眼柱镜",
    "production_date": "生产日期",
    "notes":           "备注",
    # Unique anti-counterfeiting code — auto-generated if blank, written back to Excel
    "qr_code":         "镜片码",
}

# _cache["data"]     → {order_id: order_dict}
# _cache["qr_index"] → {qr_code: order_id}
_cache = {"mtime": None, "data": {}, "qr_index": {}}


def _get_mtime():
    try:
        return os.path.getmtime(Config.EXCEL_PATH)
    except FileNotFoundError:
        return None


def _format_optical(val) -> str:
    """Normalizes optical power values to e.g. '+1.00' or '-2.50'."""
    if val is None or str(val).strip() == "":
        return ""
    try:
        f = float(val)
        sign = "+" if f >= 0 else ""
        return f"{sign}{f:.2f}"
    except (ValueError, TypeError):
        return str(val)


def load_orders() -> dict:
    """Returns dict {order_id: order_dict}. Invalidates cache by file mtime."""
    mtime = _get_mtime()
    if mtime is not None and mtime == _cache["mtime"]:
        return _cache["data"]

    if mtime is None:
        return {}

    wb = openpyxl.load_workbook(Config.EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    reverse_map = {v: k for k, v in COLUMN_MAP.items()}
    col_index = {}
    for idx, header in enumerate(headers):
        if header in reverse_map:
            col_index[reverse_map[header]] = idx

    orders = {}
    qr_index = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        order = {}
        for key, idx in col_index.items():
            val = row[idx] if idx < len(row) else None
            if key in ("left_sph", "left_cyl", "right_sph", "right_cyl"):
                order[key] = _format_optical(val)
            elif key == "production_date" and hasattr(val, "strftime"):
                order[key] = val.strftime("%Y-%m-%d")
            else:
                order[key] = str(val).strip() if val is not None else ""

        oid = order.get("order_id", "").strip()
        if not oid:
            continue
        orders[oid] = order

        qc = order.get("qr_code", "")
        if qc:
            qr_index[qc] = oid

    wb.close()
    _cache["mtime"] = mtime
    _cache["data"] = orders
    _cache["qr_index"] = qr_index
    return orders


def get_order_by_qr(qr_code: str) -> dict | None:
    """Look up an order by its unique QR code (镜片码)."""
    load_orders()
    oid = _cache["qr_index"].get(str(qr_code).strip())
    if oid is None:
        return None
    return _cache["data"].get(oid)


def assign_qr_codes() -> int:
    """
    Opens the Excel file, assigns a unique 镜片码 to any row that lacks one,
    saves the file, and invalidates the cache.
    Returns the number of new codes written.
    """
    if not os.path.exists(Config.EXCEL_PATH):
        return 0

    wb = openpyxl.load_workbook(Config.EXCEL_PATH, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Find or append the 镜片码 column
    qr_col = None
    for idx, h in enumerate(headers, start=1):
        if h == "镜片码":
            qr_col = idx
            break
    if qr_col is None:
        qr_col = len(headers) + 1
        ws.cell(row=1, column=qr_col, value="镜片码")

    # Find the 订单号 column to detect blank rows
    order_id_col = next((i + 1 for i, h in enumerate(headers) if h == "订单号"), None)

    written = 0
    for row_idx in range(2, ws.max_row + 1):
        if order_id_col and not ws.cell(row=row_idx, column=order_id_col).value:
            continue
        cell = ws.cell(row=row_idx, column=qr_col)
        if not cell.value:
            # 16-char uppercase hex — unique, hard to guess, URL-safe
            cell.value = uuid.uuid4().hex[:16].upper()
            written += 1

    wb.save(Config.EXCEL_PATH)
    wb.close()
    _cache["mtime"] = None  # invalidate cache
    return written
